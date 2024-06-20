import os
import openpyxl
from django.http import HttpResponse
import mimetypes
from openpyxl.utils import get_column_letter
from apps.catalogo.models import Orden, PersonaSistema, FormatoCliente


def createReportCustomerDinamic(request):
    idUser = request.user.pk
    objPersona = PersonaSistema.objects.get(usuario_id=idUser)
    listOrdenes = Orden.objects.filter(cliente_id=objPersona.cliente.pk).order_by("-pk")
    cadenaIdentificador = request.GET.get("idReport")
    cadenaIdentificador = cadenaIdentificador[5 : len(cadenaIdentificador)]
    idReporte = cadenaIdentificador[0:-5]
    objReporte = FormatoCliente.objects.get(pk=idReporte)

    # Crear un nuevo libro de trabajo
    workbook = openpyxl.Workbook()

    # Seleccionar una hoja de trabajo
    sheet = workbook.active

    # Agregar los encabezados dinámicos del documento
    contadorHeaders = 1
    for header in objReporte.get_campos():
        sheet.cell(1, contadorHeaders, header)
        contadorHeaders += 1

    columnNumber = 2

    # Recorrer el array y escribir los datos en el archivo Excel
    for orden in listOrdenes:
        contadorHeaders = 1
        for header in objReporte.get_campos_acceso():
            if header == "eq_type":
                sheet.cell(columnNumber, contadorHeaders, orden.get_eq_type_display())
            if header == "proveedor":
                sheet.cell(columnNumber, contadorHeaders, orden.proveedor.nombre)
            if header == "direccion_origen":
                sheet.cell(columnNumber, contadorHeaders, orden.get_origin())
            if header == "direccion_destino":
                sheet.cell(columnNumber, contadorHeaders, orden.get_destination())
            if header == "tipo_cambio":
                sheet.cell(
                    columnNumber, contadorHeaders, orden.get_tipo_cambio_display()
                )
            contadorHeaders += 1
        columnNumber += 1

    # Iterar sobre todas las celdas y obtener la longitud del valor más largo en cada columna
    column_widths = []
    max_row = len(listOrdenes) + 1
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    # Establecer el ancho de la columna en función de la longitud del valor más largo en esa columna
    for i, column_width in enumerate(column_widths):
        column_letter = get_column_letter(i + 1)
        sheet.column_dimensions[column_letter].width = column_width

    nombre_archivo = f"{objReporte.nombre}.xlsx"
    workbook.save(nombre_archivo)

    # Hacer el proceso de descarga
    with open(nombre_archivo, "rb") as fh:
        mime_type, _ = mimetypes.guess_type(nombre_archivo)
        response = HttpResponse(fh.read(), content_type=mime_type)
        response["Content-Disposition"] = "inline; filename=" + os.path.basename(
            nombre_archivo
        )
        return response
