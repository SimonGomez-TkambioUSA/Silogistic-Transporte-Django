from apps.catalogo.models import Orden
import os
from transporte import settings
from apps.catalogo.functions_reports import render_to_pdf
import mimetypes
from django.http import HttpResponse, JsonResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.db import connection


def createPDFOrder(request):
    # Crearemos la ruta por en la cual se almacenara el archivo
    orden_id = request.GET.get("orden")
    orden = Orden.objects.get(pk=orden_id)
    path_file = os.path.join(f"reportes/ordenes/")
    if not os.path.exists(f"{settings.MEDIA_ROOT}{path_file}"):
        os.makedirs(f"{settings.MEDIA_ROOT}{path_file}")
    name_documento = "REPORT_ORD-000" + str(orden_id)
    pdf_file = "{}{}.pdf".format(path_file, name_documento)
    print(orden.get_producto())
    with open(f"{settings.MEDIA_ROOT}{pdf_file}", "wb") as f:
        pdf = render_to_pdf(
            "formato_orden.html",
            {
                "numero_orden": "ORD-000" + str(orden_id),
                "equipment_orden": orden.get_eq_type_display(),
                "realase_orden": orden.realase,
                "proveedor": orden.proveedor,
                "carrier_orden": orden.get_carrier(),
                "shipment_date": orden.get_shippdate(),
                "points": orden.get_producto(),
                "delivery_orden": orden.get_delivery(),
                "delivery_date": orden.get_deliverydate(),
                "total": orden.costo_proveedor,
            },
        )
        f.write(pdf)

    # Haremos el proceso de descarga
    with open(f"{settings.MEDIA_ROOT}{pdf_file}", "rb") as fh:
        mime_type, _ = mimetypes.guess_type(f"{settings.MEDIA_ROOT}{pdf_file}")
        response = HttpResponse(fh.read(), content_type=mime_type)
        response["Content-Disposition"] = "inline; filename=" + os.path.basename(
            f"{settings.MEDIA_ROOT}{pdf_file}"
        )
        return response


def createReporteContabilidad(request):
    startDate = request.GET.get("startDate")
    endDate = request.GET.get("endDate")
    listOrdenes = Orden.objects.all().exclude(status_po_id=17).order_by("-pk")
    if startDate and endDate:
        listOrdenes = listOrdenes.filter(created_at__gte=startDate)
        listOrdenes = listOrdenes.filter(created_at__lte=endDate)
    elif startDate:
        listOrdenes = listOrdenes.filter(created_at__gte=startDate)
    elif endDate:
        listOrdenes = listOrdenes.filter(created_at__lte=endDate)

    # Crear un nuevo libro de trabajo
    workbook = openpyxl.Workbook()

    # Seleccionar una hoja de trabajo
    sheet = workbook.active

    # Agregaremos los encabezados del documento
    sheet.cell(1, 1, "NO. ORDEN")
    sheet.cell(1, 2, "PROVEEDOR")
    sheet.cell(1, 3, "ORIGEN")
    sheet.cell(1, 4, "DESTINO")
    sheet.cell(1, 5, "COSTO PROVEEDOR")
    sheet.cell(1, 6, "EQ TYPE")
    sheet.cell(1, 7, "FECHA ORDEN")

    columnNumber = 2

    # Recorrer el array y escribir los datos en el archivo Excel

    for orden in listOrdenes:
        sheet.cell(columnNumber, 1, "ORD-000" + str(orden.id))
        sheet.cell(columnNumber, 2, orden.proveedor.nombre)
        sheet.cell(columnNumber, 3, orden.get_origin())
        sheet.cell(columnNumber, 4, orden.get_destination())
        sheet.cell(columnNumber, 5, orden.costo_proveedor)
        sheet.cell(columnNumber, 6, orden.get_eq_type_display())
        sheet.cell(columnNumber, 7, orden.created_at.strftime("%d/%m/%Y"))
        columnNumber += 1

    # Itera sobre todas las celdas y obtén la longitud del valor más largo en cada columna
    column_widths = []
    max_row = len(listOrdenes) + 1
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))

    # Establece el ancho de la columna en función de la longitud del valor más largo en esa columna
    for i, column_width in enumerate(column_widths):
        column_letter = get_column_letter(i + 1)
        sheet.column_dimensions[column_letter].width = column_width

    workbook.save("Tarifas.xlsx")
    # Haremos el proceso de descarga
    with open("Tarifas.xlsx", "rb") as fh:
        mime_type, _ = mimetypes.guess_type("Tarifas.xlsx")
        response = HttpResponse(fh.read(), content_type=mime_type)
        response["Content-Disposition"] = "inline; filename=" + os.path.basename(
            "Tarifas.xlsx"
        )
        return response


def DownloadExcellOrdenes(request):
    # Obtenemos todas las ordenes dependiendo del filtro
    conection = connection.cursor()
    if request.GET.get("ordenes"):
        if request.GET.get("ordenes") == "100":
            nameDocument = "OrdenesSinFactura"
            query_bd = "select getOrdenesAllSF()"
            conection.execute(query_bd)
            query_success = dictfetchall(conection)
            ordenesAll = query_success[0]["getordenesallsf"]
        else:
            nameDocument = "OrdenesCompletas"
            query_bd = "select getOrdenesAll('{}')".format(request.GET.get("ordenes"))
            conection.execute(query_bd)
            query_success = dictfetchall(conection)
            ordenesAll = query_success[0]["getordenesall"]
    else:
        nameDocument = "OrdenesAll"
        query_bd = "select getOrdenesAll(null)"
        conection.execute(query_bd)
        query_success = dictfetchall(conection)
        ordenesAll = query_success[0]["getordenesall"]

    # Una vez que obtenemos las ordenes, agregamos las cabeceras del excell para despues hacer el recorrido e ir apendando orden por orden
    datos = [
        [
            "ORDEN",
            "W/O #",
            "CARRIER",
            "CUSTOMER",
            "SHIPMENT STATUS",
            "FACTURA SIGO",
            "FECHA FACTURA SIGO",
            "FECHA PAGO SIGO",
            "CANTIDAD PAGO SIGO",
            "FECHA DEPOSITO",
            "MONTO",
            "SALDO",
            "FECHA PAGO",
            "MONTO PAGO",
        ]
    ]

    for orden in ordenesAll:

        idOrden = orden["id"]
        datos.append(
            [
                f"ORD-000{idOrden}",
                orden["wo"],
                orden["proveedor"],
                orden["cliente"],
                orden["status_orden"],
                orden["cliente_factura"],
                orden["cliente_factura_fecha"],
                orden["cliente_pago_fecha"],
                orden["cliente_cantidad"],
                orden["factoraje_fecha_deposito"],
                orden["factoraje_monto"],
                orden["saldo"],
                orden["fecha_pago"],
                orden["monto_pago"],
            ]
        )

    # Crear un nuevo libro (archivo Excel)
    libro = openpyxl.Workbook()

    # Seleccionar la hoja activa (por defecto es la primera hoja creada)
    hoja = libro.active

    # Agregar datos a la hoja
    for fila in datos:
        hoja.append(fila)

    # Proceso para asignar el nombre del archivo

    # Guardar el archivo Excel
    libro.save(f"{nameDocument}.xlsx")

    # Haremos el proceso de descarga
    with open(f"{nameDocument}.xlsx", "rb") as fh:
        mime_type, _ = mimetypes.guess_type(f"{nameDocument}.xlsx")
        response = HttpResponse(fh.read(), content_type=mime_type)
        response["Content-Disposition"] = "inline; filename=" + os.path.basename(
            f"{nameDocument}.xlsx"
        )
        return response


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]
